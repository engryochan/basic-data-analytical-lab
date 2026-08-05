import csv
import sys
from openpyxl import load_workbook

SRC = "【英博体育_电子】7-6-7-13_结算_全阶段_3576.xlsx"

wb = load_workbook(SRC, read_only=True, data_only=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    out_path = f"{sheet_name}.csv"
    print(f"Converting sheet '{sheet_name}' -> {out_path} ...")
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        row_count = 0
        for row in ws.iter_rows(values_only=True):
            writer.writerow(row)
            row_count += 1
            if row_count % 200000 == 0:
                print(f"  ...{row_count} rows written")
        print(f"  Done: {row_count} rows total")

wb.close()
print("All sheets converted.")
